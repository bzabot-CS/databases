import sqlite3
from openpyxl import load_workbook
#import re


def sanitize_text_input(input):
    # Remove any character that is not alphanumeric or a space
    #sanitized = re.sub(r'[^a-zA-Z0-9\s]', '', s)
    """
    Acho que pelo fato de ter artigos que usam algum tipo de pontuação não da para usar esse regex, tem que pensar em outro, se for sanitizar
    """
    if (not(input)):
        return False

    # Verificar se o valor é uma string
    if isinstance(input, str):
        # Remove extra spaces at the beginning and end
        sanitized = input.strip()

        # Replace multiple spaces with a single space
        sanitized = ' '.join(sanitized.split())
        return sanitized
    
    return input

def check_existence(cur, table, column, value):
    value = sanitize_text_input(value)
    query = f"SELECT * FROM {table} WHERE {column} = ?"
    cur.execute(query, (value,))
    return cur.fetchone()

def check_existence_two_values(cur, table, column1, column2, value1, value2):
    value1 = sanitize_text_input(value1)
    value2 = sanitize_text_input(value2)

    query = f"SELECT * FROM {table} WHERE {column1} = ? AND {column2} = ? "
    cur.execute(query, (value1, value2))
    return cur.fetchone()

def table_one_value(cur, table, column, value):
    if not table.isidentifier() or not column.isidentifier():
        raise ValueError("Nomes de tabela ou coluna inválidos")
    
    value = sanitize_text_input(value)
    found_row = check_existence(cur, table, column, value)

    if(found_row): return found_row[0] #retorna o id
    elif(value):
        return insert_one_value(cur, table, column, value)

def insert_one_value(cur, table, column, value):
    query = f"""
            INSERT INTO {table} ({column}) 
            VALUES (?);
        """
    cur.execute(query, (value,))
    return cur.lastrowid

def table_two_values(cur, table, column1, column2, values, check_function = check_existence):

    if not table.isidentifier() or not column1.isidentifier() or not column2.isidentifier():
        raise ValueError("Nomes de tabela ou coluna inválidos")
    
    values = [sanitize_text_input(values[0]), sanitize_text_input(values[1])]
    if (check_function == check_existence):
        found_row = check_function(cur, table, column1, values[0])
    else:
        found_row = check_function(cur, table, column1, column2, values[0], values[1])

    if(found_row): return found_row[0] 
    elif(values[0] and values[1]):
        return insert_two_values(cur, table, column1, column2, values[0], values[1])
    
def insert_two_values(cur, table, column1, column2, value1, value2):
    query = f"""
        INSERT INTO {table} ({column1}, {column2}) 
        VALUES (?, ?);
    """
    cur.execute(query, (value1, value2))
    return cur.lastrowid

def table_local(cur, value):
    places = value.split("|")
    place = [p.split(",") for p in places]
    municipios = []
    paises = []
    for p in place:
        pais_id = table_one_value(cur, "paises", "nome", p[0])

        if(len(p) == 1): # se só informar o pais 
            if pais_id not in paises:
                paises.append(pais_id)
        
        if(len(p) == 3):
            distrito_id = table_two_values(cur, "distritos", "nome", "pais", [p[1], pais_id])
            municipio_id = table_two_values(cur, "municipios", "nome", "distrito", [p[2], distrito_id])
            if municipio_id not in municipios:
                municipios.append(municipio_id)
        #else:
            # TODO AJEITAR ESSA PARTE
            # found_row = check_existence(cur, "municipios", "nome", value)
            # print(p)
            # if(not found_row): 
            #     print(f"Municipio {p[1]} não tem distrito vinculado ")
            #     return None  
    return [municipios, paises]

def table_contrato(cur, values, ids):
        for key, value in values.items():
            values[key] = sanitize_text_input(value)
        found_row = check_existence(cur, "contratos", "idContrato", values["idcontrato"])

        if (found_row): return found_row[0]
        
        query = f"""
            INSERT INTO contratos (idContrato, objetoContrato, dataPublicacao, dataCelebracaoContrato , precoContratual, prazoExecucao, procedimentoCentralizado, tipoProcedimento, fundamentacao ) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?);
        """

        cur.execute(query, (values["idcontrato"], values["objectoContrato"], values["dataPublicacao"], values["dataCelebracaoContrato"], values["precoContratual"], values["prazoExecucao"], values["ProcedimentoCentralizado"], ids["procedimento_id"], ids["fundamentacao_id"]))
        return cur.lastrowid


# def insert

def handle_entidades(list_values, i):
    """Manipula entradas para o caso especial de 'entidades'."""
    if i != len(list_values) - 1 and len(list_values[i + 1].split(" - ", 1)) == 1:
        # Junta o próximo valor com o atual e indica que a próxima entrada deve ser pulada
        combined_value = list_values[i] + "|" + list_values[i + 1]
        return combined_value.split(" - ", 1), True
    return list_values[i].split(" - ", 1), False


def get_ids_for_multiple_values(cur, data, table_type, table, column1, column2=None, checkFunction=check_existence):
    list_values = data.split("|")
    ids = []
    skip_next = False
    for i in range(len(list_values)):
        if(table_type == table_one_value):
            contrato_id = table_one_value(cur, table, column1, list_values[i])
        elif (table_type == table_two_values and (not skip_next)):
            if (table == "classificacoesCpv"): 
                value = list_values[i].split(" - ", 1)
            elif (table == "entidades"):
                value, skip_next = handle_entidades(list_values, i)

            if (len(value) != 2): 
                print("Valor inválido ou incompleto")
                continue
            contrato_id = table_two_values(cur,table, column1, column2, value, checkFunction)

        ids.append(contrato_id)
    return ids

def new_contract(cur, row_data):
    contrato_values_id = {}

    contrato_values_id["procedimento_id"] = table_one_value(cur, "procedimentos", "descricao", row_data["tipoprocedimento"])
    contrato_values_id["tipoContrato_id"] = get_ids_for_multiple_values(cur, row_data["tipoContrato"], table_one_value, "tiposContrato", "descricao")
    contrato_values_id["fundamentacao_id"] = table_one_value(cur, "fundamentacoes", "fundamentacao", row_data["fundamentacao"])
    contrato_values_id["cpv_id"] = get_ids_for_multiple_values(cur, row_data["cpv"], table_two_values, "classificacoesCpv", "codigo", "descricao")
    [municipios_id, paises_id] = table_local(cur, row_data["localExecucao"])
    #print(row_data["adjudicante"])
    adjudicante_id = get_ids_for_multiple_values(cur, row_data["adjudicante"], table_two_values, "entidades", "nif", "designacao", check_existence_two_values)
    #if(len(adjudicante_id)> 1): print(row_data["adjudicante"])
    if ( row_data["adjudicatarios"]): 
        #if len(row_data["adjudicatarios"].split("|")) >1: print("adj "+ row_data["adjudicatarios"])
        adjudicatarios_id=get_ids_for_multiple_values(cur, row_data["adjudicatarios"], table_two_values, "entidades", "nif", "designacao", check_existence_two_values)
    else : 
        adjudicatarios_id = None
    # if len(paises_id)> 1 : 
    #     print(row_data["localExecucao"])
    #     print(paises_id)
    #     print()

    #criando contrato
    values_contrato = {"idcontrato": row_data["idcontrato"], "objectoContrato": row_data["objectoContrato"], "dataPublicacao": row_data["dataPublicacao"], "dataCelebracaoContrato": row_data["dataCelebracaoContrato"], "precoContratual": row_data["precoContratual"], "prazoExecucao": row_data["prazoExecucao"], "ProcedimentoCentralizado": row_data["ProcedimentoCentralizado"]}
    table_contrato(cur, values_contrato, contrato_values_id)


    # Fazer depois que o contrato está criado
    # Criar uma tabela para relacionar os tipos de contratos
    # Criar uma tabela para relacionar os cpvs
    # Criar uma tabela para relacionar adjudicantes e adjudicatarios


def add_dataset(sheet):
    # Conecte ao banco de dados SQLite
    con = sqlite3.connect('contratos_publicos.db')
    cur = con.cursor()

    headers = [cell.value for cell in sheet[1]] # pega a primeira linha do excel (o título das colunas), não começa com 0 pois no excel as linhas são indexadas a partir de 1 

    # Itera sobre as linhas da planilha, começando na segunda linha, pois a primeira é o cabeçalho
    for row in sheet.iter_rows(min_row=2, values_only=True):
        """
            sheet.iter_rows() - itera sobre as linhas da planilha e retorna em cada uma linha uma tupla com as celulas ou valores de cada coluna
            min_row=2 - a iteração começa na linha 2
            values_only=True - retorna só o valor da celula, não o objeto celula
        """
        row_data = dict(zip(headers, row)) #onde associamos os valores de cada coluna com o header correspondente
        """
            zip(headers, row) - combina a lista headers e row, em pares (chave, valor)
            dict - transforma essa combinação em um dicionario {chave: valor}
        """
        # TODO Melhor fazer a checagem da existencia do contrato do lado de fora pq ai não adiciona o que já foi adicionado, já que garantidamente tudo que ta no contrato esta na bd
        new_contract(cur, row_data)

    con.commit()
    print("Adicionado na BD")
    con.close()

# Carregue o arquivo Excel
workbook = load_workbook(filename='dataset/ContratosPublicos2024.xlsx')

# Selecione a planilha ativa
sheet = workbook.active

add_dataset(sheet)